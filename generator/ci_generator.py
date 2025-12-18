# -*- coding: utf-8 -*-
"""
CI Generator - Commercial Invoice Excel Generator
OT 샘플 기준 표준양식 (CI_INVOT251104-1)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from typing import List, Dict, Optional, Tuple
from datetime import datetime
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class SenderInfo:
    """SENDER 정보 (Beautyselection)"""
    company_name: str
    address: str
    tel: str = ""
    fax: str = ""
    business_no: str = ""
    email: str = ""
    phone: str = ""


# DEFAULT: SENDER 정보 (국내/해외)
SENDER_DOMESTIC = SenderInfo(
    company_name="주식회사 뷰티셀렉션",
    address="서울특별시 강남구 도산대로100길 22-5",
    business_no="551-81-01715"
)

SENDER_INTERNATIONAL = SenderInfo(
    company_name="Beautyselection Co., Ltd.",
    address="Beautyselection, 22-5, Dosan-daero 100-gil, Gangnam-gu, Seoul, Republic of Korea"
)

# Bank Details (국제 거래용)
BANK_DETAILS = {
    'account_holder': 'BeautySelection Co., Ltd.',
    'account_number': '140-015-398297',
    'bank_name': 'SHINHAN BANK',
    'bank_address': '1585, Sangam-dong, mapo-gu, Seoul, South Korea',
    'swift_code': 'HVBKKRSEXXX',
    'company_address': '31, Bongeunsa-ro 68-gil, Gangnam-gu, Seoul, Republic of Korea'
}


@dataclass
class ReceiverInfo:
    """RECEIVER 정보 (고객사)"""
    company_name: str
    address: str
    country: str = ""
    email: str = ""
    phone: str = ""
    business_no: str = ""


@dataclass
class ShippingTerms:
    """SHIPPING TERMS"""
    terms: str = ""  # FOB, CIF, EXW
    loading_port: str = ""
    destination_port: str = ""
    shipping_method: str = ""  # BY AIR, BY SEA
    reason_of_export: str = ""  # SALE, SAMPLE, FOC


@dataclass
class LineItem:
    """상품 라인 아이템"""
    sku_id: str
    barcode: str = ""
    description: str = ""
    hs_code: str = ""
    qty: int = 0
    qty_outbox: float = 0.0  # QTY(OUTBOX)
    unit_price: float = 0.0
    is_foc: bool = False
    note: str = ""


@dataclass
class InvoiceData:
    """CI 전체 데이터"""
    # HEADER
    invoice_date: datetime
    invoice_no: str
    order_no: str = ""

    # SENDER (담당자 정보)
    staff_email: str = ""
    staff_phone: str = ""
    is_domestic: bool = False

    # RECEIVER
    receiver: ReceiverInfo = None

    # SHIPPING TERMS
    shipping: ShippingTerms = None

    # CURRENCY
    currency: str = "KRW"

    # ITEMS
    items: List[LineItem] = field(default_factory=list)

    # SUMMARY
    tax_rate: float = 0.0
    total_transaction: float = 0.0

    # REMARKS
    custom_remarks: str = ""


class CIGenerator:
    """Commercial Invoice Generator - OT 샘플 기준"""

    # 스타일 정의
    TITLE_FONT = Font(bold=True, size=18)
    HEADER_FONT = Font(bold=True, size=10)
    LABEL_FONT = Font(bold=True, size=9)
    VALUE_FONT = Font(size=9)
    SMALL_FONT = Font(size=8)

    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    GRAY_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # 로고 경로 (선택)
    LOGO_PATH = Path(__file__).parent.parent / "assets" / "BS_LOGO.jpg"

    def __init__(self):
        pass

    def validate(self, data: InvoiceData) -> Tuple[bool, List[str]]:
        """필수 필드 검증"""
        errors = []

        if not data.invoice_no:
            errors.append("Invoice No. is required")
        if not data.invoice_date:
            errors.append("Invoice Date is required")
        if not data.staff_email:
            errors.append("Staff email is required")
        if not data.receiver:
            errors.append("Receiver info is required")
        elif not data.receiver.company_name:
            errors.append("Receiver company name is required")
        elif not data.receiver.address:
            errors.append("Receiver address is required")
        if data.currency not in ['KRW', 'USD', 'EUR']:
            errors.append("Currency must be KRW, USD, or EUR")
        if not data.items:
            errors.append("At least one item is required")
        else:
            for i, item in enumerate(data.items, 1):
                if not item.sku_id:
                    errors.append(f"Item {i}: SKU ID is required")
                if not item.description:
                    errors.append(f"Item {i}: Description is required")
                if item.qty <= 0:
                    errors.append(f"Item {i}: Quantity must be > 0")

        return (len(errors) == 0, errors)

    def generate(self, data: InvoiceData) -> Workbook:
        """CI Excel 생성 (OT 샘플 기준)"""
        valid, errors = self.validate(data)
        if not valid:
            raise ValueError(f"Validation failed: {'; '.join(errors)}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Commercial Invoice"

        # 컬럼 너비 설정
        col_widths = {'A': 3, 'B': 14, 'C': 14, 'D': 35, 'E': 13, 'F': 8, 'G': 10, 'H': 12, 'I': 12}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        row = 1

        # === ROW 1-2: 로고 + 타이틀 ===
        # 로고 (있으면)
        if self.LOGO_PATH.exists():
            try:
                img = Image(str(self.LOGO_PATH))
                img.width = 120
                img.height = 40
                ws.add_image(img, 'B1')
            except:
                pass

        # 타이틀
        ws.merge_cells('D1:G2')
        title_cell = ws['D1']
        title_cell.value = "COMMERCIAL INVOICE"
        title_cell.font = self.TITLE_FONT
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        row = 4

        # === SENDER + INVOICE INFO ===
        # SENDER (좌측)
        ws[f'B{row}'] = "SENDER"
        ws[f'B{row}'].font = self.LABEL_FONT

        sender = SENDER_DOMESTIC if data.is_domestic else SENDER_INTERNATIONAL
        ws[f'B{row+1}'] = sender.company_name
        ws[f'B{row+1}'].font = self.VALUE_FONT
        ws[f'B{row+2}'] = sender.address
        ws[f'B{row+2}'].font = self.VALUE_FONT
        ws[f'B{row+3}'] = data.staff_email
        ws[f'B{row+3}'].font = self.VALUE_FONT
        if data.staff_phone:
            ws[f'B{row+4}'] = data.staff_phone
            ws[f'B{row+4}'].font = self.VALUE_FONT

        # Invoice Info (우측)
        ws[f'G{row}'] = "DATE OF INVOICE"
        ws[f'G{row}'].font = self.LABEL_FONT
        ws[f'H{row}'] = data.invoice_date.strftime('%Y-%m-%d')
        ws[f'H{row}'].font = self.VALUE_FONT

        ws[f'G{row+1}'] = "INVOICE NO."
        ws[f'G{row+1}'].font = self.LABEL_FONT
        ws[f'H{row+1}'] = data.invoice_no
        ws[f'H{row+1}'].font = self.VALUE_FONT

        if data.order_no:
            ws[f'G{row+2}'] = "RELATED ORDER NO. / DATE"
            ws[f'G{row+2}'].font = self.LABEL_FONT
            ws[f'H{row+2}'] = data.order_no
            ws[f'H{row+2}'].font = self.VALUE_FONT

        row += 6

        # === RECEIVER ===
        ws[f'B{row}'] = "RECEIVER"
        ws[f'B{row}'].font = self.LABEL_FONT
        row += 1

        ws[f'B{row}'] = data.receiver.company_name
        ws[f'B{row}'].font = self.VALUE_FONT
        row += 1

        ws[f'B{row}'] = data.receiver.address
        ws[f'B{row}'].font = self.VALUE_FONT
        row += 1

        if data.receiver.email:
            ws[f'B{row}'] = data.receiver.email
            ws[f'B{row}'].font = self.VALUE_FONT
            row += 1

        if data.receiver.phone:
            ws[f'B{row}'] = data.receiver.phone
            ws[f'B{row}'].font = self.VALUE_FONT
            row += 1

        row += 2

        # === SHIPPING TERMS 테이블 (해외만) ===
        if not data.is_domestic and data.shipping:
            # 헤더 행
            shipping_headers = ['SHIPPING TERMS', 'LOADING PORT', 'DESTINATION PORT',
                              'SHIPPING METHOD', 'REASON OF EXPORT', 'CURRENCY']
            shipping_cols = ['B', 'C', 'D', 'E', 'F', 'G']

            for col, header in zip(shipping_cols, shipping_headers):
                cell = ws[f'{col}{row}']
                cell.value = header
                cell.font = self.LABEL_FONT
                cell.fill = self.GRAY_FILL
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            ws.row_dimensions[row].height = 25
            row += 1

            # 값 행
            shipping_values = [
                data.shipping.terms or '',
                data.shipping.loading_port or 'KOREA',
                data.shipping.destination_port or '',
                data.shipping.shipping_method or '',
                data.shipping.reason_of_export or 'SALE',
                data.currency
            ]

            for col, value in zip(shipping_cols, shipping_values):
                cell = ws[f'{col}{row}']
                cell.value = value
                cell.font = self.VALUE_FONT
                cell.border = self.THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')

            row += 2

        # === ITEM TABLE ===
        item_headers = ['SKU ID', 'BARCODE', 'DESCRIPTION OF GOODS', 'HS CODE',
                       'QTY\n(EA)', 'QTY\n(OUTBOX)', 'UNIT PRICE', 'TOTAL']
        item_cols = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']

        # 헤더 행
        for col, header in zip(item_cols, item_headers):
            cell = ws[f'{col}{row}']
            cell.value = header
            cell.font = self.LABEL_FONT
            cell.fill = self.GRAY_FILL
            cell.border = self.THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[row].height = 28
        row += 1

        # 아이템 행
        subtotal = 0
        total_qty = 0
        total_outbox = 0.0
        has_foc = False

        for item in data.items:
            line_total = 0 if item.is_foc else item.qty * item.unit_price
            subtotal += line_total
            total_qty += item.qty
            total_outbox += item.qty_outbox
            if item.is_foc:
                has_foc = True

            # SKU ID
            ws[f'B{row}'] = item.sku_id
            ws[f'B{row}'].font = self.VALUE_FONT
            ws[f'B{row}'].border = self.THIN_BORDER

            # BARCODE
            ws[f'C{row}'] = item.barcode
            ws[f'C{row}'].font = self.VALUE_FONT
            ws[f'C{row}'].border = self.THIN_BORDER

            # DESCRIPTION
            desc = item.description
            if item.is_foc and "- FOC" not in desc:
                desc = f"{desc} - FOC"
            ws[f'D{row}'] = desc
            ws[f'D{row}'].font = self.VALUE_FONT
            ws[f'D{row}'].border = self.THIN_BORDER
            ws[f'D{row}'].alignment = Alignment(wrap_text=True)

            # HS CODE
            ws[f'E{row}'] = item.hs_code
            ws[f'E{row}'].font = self.VALUE_FONT
            ws[f'E{row}'].border = self.THIN_BORDER

            # QTY(EA)
            ws[f'F{row}'] = f"{item.qty:,}"
            ws[f'F{row}'].font = self.VALUE_FONT
            ws[f'F{row}'].border = self.THIN_BORDER
            ws[f'F{row}'].alignment = Alignment(horizontal='right')

            # QTY(OUTBOX)
            ws[f'G{row}'] = f"{item.qty_outbox:.1f}" if item.qty_outbox else ""
            ws[f'G{row}'].font = self.VALUE_FONT
            ws[f'G{row}'].border = self.THIN_BORDER
            ws[f'G{row}'].alignment = Alignment(horizontal='right')

            # UNIT PRICE
            if item.is_foc:
                ws[f'H{row}'] = f"{data.currency} {int(item.unit_price):,}" if item.unit_price else f"{data.currency} 0"
            else:
                ws[f'H{row}'] = f"{data.currency} {int(item.unit_price):,}" if data.currency != 'USD' else f"USD {item.unit_price:,.2f}"
            ws[f'H{row}'].font = self.VALUE_FONT
            ws[f'H{row}'].border = self.THIN_BORDER
            ws[f'H{row}'].alignment = Alignment(horizontal='right')

            # TOTAL
            ws[f'I{row}'] = f"{data.currency} {int(line_total):,}" if data.currency != 'USD' else f"USD {line_total:,.2f}"
            ws[f'I{row}'].font = self.VALUE_FONT
            ws[f'I{row}'].border = self.THIN_BORDER
            ws[f'I{row}'].alignment = Alignment(horizontal='right')

            row += 1

        # TOTAL 행
        ws[f'B{row}'] = "TOTAL"
        ws[f'B{row}'].font = self.LABEL_FONT
        ws[f'B{row}'].border = self.THIN_BORDER

        for col in ['C', 'D', 'E']:
            ws[f'{col}{row}'].border = self.THIN_BORDER

        ws[f'F{row}'] = f"{total_qty:,}"
        ws[f'F{row}'].font = self.LABEL_FONT
        ws[f'F{row}'].border = self.THIN_BORDER
        ws[f'F{row}'].alignment = Alignment(horizontal='right')

        ws[f'G{row}'] = f"{total_outbox:.1f}"
        ws[f'G{row}'].font = self.LABEL_FONT
        ws[f'G{row}'].border = self.THIN_BORDER
        ws[f'G{row}'].alignment = Alignment(horizontal='right')

        ws[f'H{row}'].border = self.THIN_BORDER

        # TOTAL 금액 (노란 배경)
        ws[f'I{row}'] = f"₩{int(subtotal):,}" if data.currency == 'KRW' else f"${subtotal:,.2f}"
        ws[f'I{row}'].font = self.LABEL_FONT
        ws[f'I{row}'].border = self.THIN_BORDER
        ws[f'I{row}'].fill = self.YELLOW_FILL
        ws[f'I{row}'].alignment = Alignment(horizontal='right')

        row += 2
        remarks_start_row = row

        # === REMARKS (좌측) ===
        ws[f'B{row}'] = "REMARKS"
        ws[f'B{row}'].font = self.LABEL_FONT
        row += 1

        remarks_lines = [
            "1. This invoice is valid for one month from the date of creation.",
            "2. Payment terms: T/T 100% before shipment.",
            "3. Bank Details:",
            f"   Account Holder Name: {BANK_DETAILS['account_holder']}",
            f"   Bank Account Number: {BANK_DETAILS['account_number']}",
            f"   Name of the Bank: {BANK_DETAILS['bank_name']}",
            f"   Bank Address: {BANK_DETAILS['bank_address']}",
            f"   SWIFT CODE : {BANK_DETAILS['swift_code']}",
            f"   Company Address: {BANK_DETAILS['company_address']}"
        ]

        if data.is_domestic:
            remarks_lines = [
                "1. 본 견적은 작성일로부터 1개월간 유효합니다.",
                "2. 결제조건: 출고 전 100%",
                "3. 입금계좌:",
                f"   은행: 신한은행",
                f"   계좌번호: {BANK_DETAILS['account_number']}",
                f"   예금주: ㈜뷰티셀렉션"
            ]

        for line in remarks_lines:
            ws[f'B{row}'] = line
            ws[f'B{row}'].font = self.SMALL_FONT
            row += 1

        # FOC 문구
        if has_foc:
            row += 1
            ws[f'B{row}'] = "No commercial value but for customs purpose only."
            ws[f'B{row}'].font = self.SMALL_FONT
            row += 1

        # Custom remarks
        if data.custom_remarks:
            row += 1
            ws[f'B{row}'] = data.custom_remarks
            ws[f'B{row}'].font = self.SMALL_FONT
            row += 1

        # AUTHORIZED SIGNATURE
        row += 1
        ws[f'B{row}'] = "AUTHORIZED SIGNATURE"
        ws[f'B{row}'].font = self.LABEL_FONT

        # === SUMMARY (우측) ===
        summary_row = remarks_start_row

        # SUBTOTAL
        ws[f'G{summary_row}'] = "SUBTOTAL"
        ws[f'G{summary_row}'].font = self.LABEL_FONT
        ws[f'G{summary_row}'].border = self.THIN_BORDER
        ws[f'H{summary_row}'] = data.currency
        ws[f'H{summary_row}'].font = self.VALUE_FONT
        ws[f'H{summary_row}'].border = self.THIN_BORDER
        ws[f'I{summary_row}'] = f"{int(subtotal):,}" if data.currency != 'USD' else f"{subtotal:,.2f}"
        ws[f'I{summary_row}'].font = self.VALUE_FONT
        ws[f'I{summary_row}'].border = self.THIN_BORDER
        ws[f'I{summary_row}'].alignment = Alignment(horizontal='right')
        summary_row += 1

        # TAX
        tax_amount = subtotal * data.tax_rate
        ws[f'G{summary_row}'] = "TAX"
        ws[f'G{summary_row}'].font = self.LABEL_FONT
        ws[f'G{summary_row}'].border = self.THIN_BORDER
        ws[f'H{summary_row}'] = data.currency
        ws[f'H{summary_row}'].font = self.VALUE_FONT
        ws[f'H{summary_row}'].border = self.THIN_BORDER
        ws[f'I{summary_row}'] = f"{int(tax_amount):,}" if data.currency != 'USD' else f"{tax_amount:,.2f}"
        ws[f'I{summary_row}'].font = self.VALUE_FONT
        ws[f'I{summary_row}'].border = self.THIN_BORDER
        ws[f'I{summary_row}'].alignment = Alignment(horizontal='right')
        summary_row += 1

        # TOTAL (Declaration)
        total_declaration = subtotal + tax_amount
        ws[f'G{summary_row}'] = "TOTAL (Declaration)"
        ws[f'G{summary_row}'].font = self.LABEL_FONT
        ws[f'G{summary_row}'].border = self.THIN_BORDER
        ws[f'G{summary_row}'].fill = self.YELLOW_FILL
        ws[f'H{summary_row}'] = data.currency
        ws[f'H{summary_row}'].font = self.VALUE_FONT
        ws[f'H{summary_row}'].border = self.THIN_BORDER
        ws[f'H{summary_row}'].fill = self.YELLOW_FILL
        ws[f'I{summary_row}'] = f"{int(total_declaration):,}" if data.currency != 'USD' else f"{total_declaration:,.2f}"
        ws[f'I{summary_row}'].font = self.LABEL_FONT
        ws[f'I{summary_row}'].border = self.THIN_BORDER
        ws[f'I{summary_row}'].fill = self.YELLOW_FILL
        ws[f'I{summary_row}'].alignment = Alignment(horizontal='right')
        summary_row += 1

        # TOTAL (Transaction) - FOC인 경우
        if has_foc:
            ws[f'G{summary_row}'] = "TOTAL (Transaction)"
            ws[f'G{summary_row}'].font = self.LABEL_FONT
            ws[f'G{summary_row}'].border = self.THIN_BORDER
            ws[f'G{summary_row}'].fill = self.YELLOW_FILL
            ws[f'H{summary_row}'] = data.currency
            ws[f'H{summary_row}'].font = self.VALUE_FONT
            ws[f'H{summary_row}'].border = self.THIN_BORDER
            ws[f'H{summary_row}'].fill = self.YELLOW_FILL
            ws[f'I{summary_row}'] = f"{int(data.total_transaction):,}" if data.currency != 'USD' else f"{data.total_transaction:,.2f}"
            ws[f'I{summary_row}'].font = self.LABEL_FONT
            ws[f'I{summary_row}'].border = self.THIN_BORDER
            ws[f'I{summary_row}'].fill = self.YELLOW_FILL
            ws[f'I{summary_row}'].alignment = Alignment(horizontal='right')

        return wb

    def save(self, wb: Workbook, output_path: str):
        """Workbook 저장"""
        wb.save(output_path)

    @staticmethod
    def generate_invoice_number(customer_code: str, date: datetime, sequence: int = 1) -> str:
        """Invoice 번호 생성: INV{customer_code}{YYMMDD}-{sequence}"""
        date_str = date.strftime('%y%m%d')
        return f"INV{customer_code.upper()}{date_str}-{sequence}"


if __name__ == "__main__":
    # 테스트 (OT 샘플 재현)
    generator = CIGenerator()

    test_data = InvoiceData(
        invoice_date=datetime(2025, 11, 4),
        invoice_no="INVOT251104-1",
        order_no="PO25001509",
        staff_email="jungbin.park@beautyselection.co.kr",
        staff_phone="+82-10-9443-5432",
        is_domestic=False,
        receiver=ReceiverInfo(
            company_name="Orien Trade OÜ",
            address="Puiestee 2, Tartu 50303, Estonia",
            email="nga.vu@orientrade.com",
            phone="+372 5674 8016"
        ),
        shipping=ShippingTerms(
            terms="FOB",
            loading_port="KOREA",
            destination_port="ESTONIA",
            shipping_method="BY AIR",
            reason_of_export="SALE"
        ),
        currency="KRW",
        items=[
            LineItem(
                sku_id="BIO-S023027889",
                barcode="8809891185139",
                description="[Biodance] Collagen Gel Toner Pads 10Pads",
                hs_code="3307.90-9000",
                qty=4000,
                qty_outbox=40.0,
                unit_price=10,
                is_foc=True
            )
        ],
        total_transaction=0
    )

    # 출력 폴더 생성
    output_dir = Path(__file__).parent.parent / "data" / "generated"
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = generator.generate(test_data)
    output_path = output_dir / "test_ci_ot_style.xlsx"
    generator.save(wb, str(output_path))
    print(f"Generated: {output_path}")
